import streamlit as st
import re
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ---------------------------------------
# SAFE AMOUNT PARSER
# ---------------------------------------
def parse_amount(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    s = s.replace(",", "").replace("\u00A0", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    s = re.sub(r"[^\d\.\-]", "", s)
    if not s:
        return 0.0
    try:
        return float(s)
    except:
        return 0.0


# ---------------------------------------
# FIND LAST USED COLUMN
# ---------------------------------------
def find_last_used_column(ws):
    last = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                last = max(last, cell.column)
    return last


# ---------------------------------------
# UNIVERSAL STRUCTURE DETECTION
# ---------------------------------------
def detect_structure(ws):
    max_col = ws.max_column
    max_row = ws.max_row

    # ------------------------------------------------
    # 1️⃣ DETECT UNIT COLUMN BASED ON FILE TYPE
    # ------------------------------------------------
    row1_val = ws.cell(row=1, column=1).value
    row1_text = str(row1_val).strip().lower() if row1_val else ""

    unit_col = None  # 0-based

    # Rule 1: Affordable → Column C
    if row1_text.startswith("affordable"):
        unit_col = 2  # column C (0-based)

    # Rule 2: Rent Roll → Column A
    elif row1_text.startswith("rent"):
        unit_col = 0  # column A (0-based)

    if unit_col is None:
        raise Exception("Unit column could not be detected (Row 1 mismatch).")

    # ------------------------------------------------
    # 2️⃣ DETECT RENT CODE COLUMN (original + fallback)
    # ------------------------------------------------
    code_col = None

    # Original logic: row 6
    for c in range(1, max_col + 1):
        v = ws.cell(row=6, column=c).value
        if isinstance(v, str) and v.strip().lower() in ["code", "rent code"]:
            code_col = c - 1
            break

    # Fallback: rows 7–12
    if code_col is None:
        for rr in range(7, 13):
            for c in range(1, max_col + 1):
                v = ws.cell(row=rr, column=c).value
                if isinstance(v, str) and v.strip().lower() in ["code", "rent code"]:
                    code_col = c - 1
                    break
            if code_col is not None:
                break

    if code_col is None:
        raise Exception("Rent Code column not found in row 6 or rows 7–12.")

    # ------------------------------------------------
    # 3️⃣ DETECT AMOUNT COLUMN
    # ------------------------------------------------
    amount_col = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=6, column=c).value
        if isinstance(v, str) and "amount" in v.strip().lower():
            amount_col = c - 1
            break

    if amount_col is None:
        amount_col = code_col + 1

    # ------------------------------------------------
    # 4️⃣ DETECT NAME COLUMN
    # ------------------------------------------------
    name_col = None

    # row 6
    for c in range(1, max_col + 1):
        v = ws.cell(row=6, column=c).value
        if isinstance(v, str) and "name" in v.strip().lower():
            name_col = c - 1
            break

    # row 5
    if name_col is None:
        for c in range(1, max_col + 1):
            v = ws.cell(row=5, column=c).value
            if isinstance(v, str) and "name" in v.strip().lower():
                name_col = c - 1
                break

    if name_col is None:
        st.warning("⚠ Name column not found (row 6/5). Using blank.")
        name_col = 9999

    return {
        "unit_col": unit_col,
        "code_col": code_col,
        "amount_col": amount_col,
        "name_col": name_col
    }


# ---------------------------------------
# EXTRACTION LOGIC
# ---------------------------------------
def extract_rentroll_from_ws(ws):
    info = detect_structure(ws)

    unit_col = info["unit_col"]
    code_col = info["code_col"]
    name_col = info["name_col"]
    amount_col = info["amount_col"]

    max_row = ws.max_row
    max_col = ws.max_column

    units = []
    current = None
    codes_seen = []
    codes_set = set()

    # ---------------------------------------
    # GET UNIT VALUES (non-bold only)
    # ---------------------------------------
    unit_values = set()
    col_excel = unit_col + 1

    for r in range(1, max_row + 1):
        cell = ws.cell(row=r, column=col_excel)

        # skip bold headers
        if cell.font and cell.font.bold:
            continue

        value = cell.value
        if value and str(value).strip():
            unit_values.add(str(value).strip())

    # ---------------------------------------
    # MAIN EXTRACTION LOOP
    # ---------------------------------------
    for r in range(7, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        def val(i): return row_vals[i] if i < len(row_vals) else None

        unit_raw = val(unit_col)
        code_raw = val(code_col)
        name_raw = val(name_col)
        amount_val = parse_amount(val(amount_col))

        # Start of new unit block
        if unit_raw and str(unit_raw).strip() in unit_values:
            if current:
                units.append(current)
            current = {
                "unit": str(unit_raw).strip(),
                "name": name_raw if isinstance(name_raw, str) else "",
                "charges": {},
                "total": 0.0
            }

        if current is None:
            continue

        # Total line
        if isinstance(code_raw, str) and code_raw.strip().lower() == "total":
            current["total"] = amount_val
            units.append(current)
            current = None
            continue

        # Charge entries
        if code_raw:
            code_key = str(code_raw).strip().lower()
            current["charges"][code_key] = current["charges"].get(code_key, 0.0) + amount_val
            if code_key not in codes_set:
                codes_set.add(code_key)
                codes_seen.append(code_key)

    if current:
        units.append(current)

    return units, codes_seen


# ---------------------------------------
# APPEND TO ORIGINAL
# ---------------------------------------
def append_extracted_to_original(ws, units, codes_seen):
    max_col = find_last_used_column(ws)
    start_col = max_col + 1

    ws.cell(row=1, column=start_col, value="Resident Name")
    col_map = {"name": start_col}

    c = start_col + 1
    for code in codes_seen:
        ws.cell(row=1, column=c, value=code)
        col_map[code] = c
        c += 1

    ws.cell(row=1, column=c, value="Total Amount")
    col_map["total"] = c

    unit_col_index = detect_structure(ws)["unit_col"] + 1
    unit_rows = {}

    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=unit_col_index).value
        if v:
            unit_rows[str(v).strip()] = r

    for u in units:
        if u["unit"] not in unit_rows:
            continue
        r = unit_rows[u["unit"]]

        ws.cell(row=r, column=col_map["name"], value=u["name"])
        for code in codes_seen:
            ws.cell(row=r, column=col_map[code], value=u["charges"].get(code, 0.0))
        ws.cell(row=r, column=col_map["total"], value=u["total"])

    return list(col_map.values())


# ---------------------------------------
# AUTO-FIT
# ---------------------------------------
def autofit_specific_columns(ws, cols):
    for col in cols:
        max_len = 0
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=col).value
            if val:
                s = str(val).strip()
                max_len = max(max_len, len(s))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2


# ---------------------------------------
# FORMAT NEW COLUMNS
# ---------------------------------------
def format_new_columns(ws, cols):
    style = Font(bold=True, color="E20000")
    for col in cols:
        for r in range(1, ws.max_row + 1):
            c = ws.cell(row=r, column=col)
            if c.value is not None:
                c.font = style


# ---------------------------------------
# STREAMLIT UI
# ---------------------------------------
st.title("🏢 Rent Charge Codes Extractor \n * Charge Codes Extractor From Rent Roll or Affordable Rent Roll Excel File.")

uploaded = st.file_uploader("Ensure to upload the file for Rent Roll or Affordable Rent Roll Only.", type=["xlsx"])

if uploaded:
    st.success("Uploaded!")

    if st.button("Start Extracting"):
        try:
            progress = st.progress(0)
            status = st.empty()

            wb = load_workbook(uploaded, data_only=True)
            ws = wb.active

            status.write("🔍 Detecting structure...")
            progress.progress(25)

            status.write("📄 Extracting charges...")
            units, codes = extract_rentroll_from_ws(ws)
            progress.progress(50)

            status.write("📝 Appending extracted data...")
            new_cols = append_extracted_to_original(ws, units, codes)
            progress.progress(75)

            status.write("📏 Auto-adjusting column widths...")
            autofit_specific_columns(ws, new_cols)
            progress.progress(90)

            status.write("🎨 Highlighting updated cells...")
            format_new_columns(ws, new_cols)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            progress.progress(100)

            st.success("✅ Processing completed!")
            st.download_button(
                "⬇ Download Processed Excel File",
                buf,
                f"processed_{uploaded.name}",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"❌ Error: {e}")

